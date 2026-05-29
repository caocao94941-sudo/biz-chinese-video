#!/usr/bin/env python3
"""Web application for biz-chinese-video generator."""

import asyncio
import io
import json
import os
import sqlite3
import uuid
from datetime import date
from pathlib import Path

from flask import (
    Flask, render_template, request, jsonify,
    send_file, send_from_directory, redirect, url_for
)
from openpyxl import load_workbook

import sys
sys.path.insert(0, str(Path(__file__).resolve().parent))

from scripts.utils.db_helper import (
    DB_PATH, TEMPLATES_DIR, OUTPUT_DIR, ASSETS_DIR,
    get_connection, init_db, get_full_lesson_data,
    get_lesson, get_vocabulary, get_grammar,
)

app = Flask(
    __name__,
    template_folder=str(Path(__file__).parent / "web" / "templates"),
    static_folder=str(Path(__file__).parent / "web" / "static"),
)
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16MB

# Ensure DB exists
if not DB_PATH.exists():
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    init_db()


def import_lesson_from_json(data: dict) -> int:
    """Import a lesson from JSON template into DB. Returns lesson_id."""
    conn = get_connection()
    lesson = data["lesson"]
    slug = lesson.get("title_en", "lesson").lower().replace(" ", "_")
    slug = f"{slug}_{uuid.uuid4().hex[:6]}"

    conn.execute(
        """INSERT INTO lessons (slug, title_zh, title_pinyin, title_en,
           hsk_level, category, topic_image, business_tip_zh, business_tip_en)
           VALUES (?,?,?,?,?,?,?,?,?)""",
        (slug, lesson["title_zh"], lesson["title_pinyin"], lesson["title_en"],
         lesson.get("hsk_level", "HSK1"), lesson.get("category", "商务汉语"),
         None, lesson.get("business_tip_zh", ""), lesson.get("business_tip_en", "")),
    )
    lesson_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    for i, s in enumerate(data.get("sentences", []), 1):
        hl = json.dumps(s.get("highlight_words", []), ensure_ascii=False)
        conn.execute(
            """INSERT INTO sentences (lesson_id, sort_order, text_zh, text_pinyin,
               text_en, highlight_words, duration_ms) VALUES (?,?,?,?,?,?,?)""",
            (lesson_id, i, s["text_zh"], s["text_pinyin"], s["text_en"], hl, 3000),
        )

    for i, v in enumerate(data.get("vocabulary", []), 1):
        conn.execute(
            """INSERT INTO vocabulary (lesson_id, sort_order, word_zh, word_pinyin,
               word_en, word_pos, hsk_level, icon_emoji, first_appear_sentence)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (lesson_id, i, v["word_zh"], v["word_pinyin"], v["word_en"],
             v.get("word_pos", ""), v.get("hsk_level", "HSK1"),
             v.get("icon_emoji", ""), v.get("first_appear_sentence", 1)),
        )

    for i, g in enumerate(data.get("grammar", []), 1):
        conn.execute(
            """INSERT INTO grammar (lesson_id, sort_order, pattern_zh, pattern_pinyin,
               pattern_en, explanation_zh, explanation_en, example_zh, example_pinyin,
               example_en, hsk_level, first_appear_sentence)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (lesson_id, i, g["pattern_zh"], g.get("pattern_pinyin", ""),
             g.get("pattern_en", ""), g.get("explanation_zh", ""),
             g.get("explanation_en", ""), g.get("example_zh", ""),
             g.get("example_pinyin", ""), g.get("example_en", ""),
             g.get("hsk_level", "HSK1"), g.get("first_appear_sentence", 1)),
        )

    conn.commit()
    conn.close()
    return lesson_id


def parse_excel_to_dict(file_bytes: bytes) -> dict:
    """Parse uploaded Excel file into lesson dict (same format as JSON template)."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)

    # Sheet 1: 课程信息
    ws1 = wb["课程信息"]
    info = {}
    for row in ws1.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0] and row[1]:
            info[row[0].strip()] = str(row[1]).strip()

    lesson = {
        "title_zh": info.get("课程标题（中文）", ""),
        "title_pinyin": info.get("课程标题（拼音）", ""),
        "title_en": info.get("课程标题（英文）", ""),
        "hsk_level": info.get("HSK等级", "HSK1"),
        "category": info.get("分类", "商务汉语"),
        "business_tip_zh": info.get("商务小贴士（中文）", ""),
        "business_tip_en": info.get("商务小贴士（英文）", ""),
    }
    settings = {
        "voice": info.get("语音", "zh-CN-XiaoxiaoNeural"),
        "video_format": info.get("视频格式", "both"),
    }

    # Sheet 2: 课文句子
    ws2 = wb["课文句子"]
    sentences = []
    for row in ws2.iter_rows(min_row=2, max_col=5, values_only=True):
        if not row[1]:
            continue
        hl = [w.strip() for w in str(row[4] or "").split(",") if w.strip()]
        sentences.append({
            "text_zh": str(row[1]).strip(),
            "text_pinyin": str(row[2] or "").strip(),
            "text_en": str(row[3] or "").strip(),
            "highlight_words": hl,
        })

    # Sheet 3: 词汇
    ws3 = wb["词汇"]
    vocabulary = []
    for row in ws3.iter_rows(min_row=2, max_col=7, values_only=True):
        if not row[0]:
            continue
        vocabulary.append({
            "word_zh": str(row[0]).strip(),
            "word_pinyin": str(row[1] or "").strip(),
            "word_en": str(row[2] or "").strip(),
            "word_pos": str(row[3] or "").strip(),
            "hsk_level": str(row[4] or "HSK1").strip(),
            "icon_emoji": str(row[5] or "").strip(),
            "first_appear_sentence": int(row[6] or 1),
        })

    # Sheet 4: 语法
    ws4 = wb["语法"]
    grammar = []
    for row in ws4.iter_rows(min_row=2, max_col=10, values_only=True):
        if not row[0]:
            continue
        grammar.append({
            "pattern_zh": str(row[0]).strip(),
            "pattern_pinyin": str(row[1] or "").strip(),
            "pattern_en": str(row[2] or "").strip(),
            "explanation_zh": str(row[3] or "").strip(),
            "explanation_en": str(row[4] or "").strip(),
            "example_zh": str(row[5] or "").strip(),
            "example_pinyin": str(row[6] or "").strip(),
            "example_en": str(row[7] or "").strip(),
            "hsk_level": str(row[8] or "HSK1").strip(),
            "first_appear_sentence": int(row[9] or 1),
        })

    wb.close()
    return {
        "lesson": lesson,
        "sentences": sentences,
        "vocabulary": vocabulary,
        "grammar": grammar,
        "settings": settings,
    }


# --- Routes ---

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/template/download")
def download_template():
    """Download the Excel lesson template."""
    template_path = TEMPLATES_DIR / "lesson_template.xlsx"
    if not template_path.exists():
        # Generate if missing
        from scripts.create_excel_template import create_template
        create_template(template_path, with_sample=False)
    return send_file(str(template_path), as_attachment=True,
                     download_name="lesson_template.xlsx")


@app.route("/api/template/sample")
def download_sample():
    """Download the sample Excel with data."""
    sample_path = TEMPLATES_DIR / "lesson_sample.xlsx"
    if not sample_path.exists():
        from scripts.create_excel_template import create_template
        create_template(sample_path, with_sample=True)
    return send_file(str(sample_path), as_attachment=True,
                     download_name="lesson_sample.xlsx")


@app.route("/api/upload", methods=["POST"])
def upload_lesson():
    """Upload an Excel or JSON lesson file, import to DB, return lesson_id."""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    filename = file.filename.lower()

    try:
        if filename.endswith(".xlsx"):
            data = parse_excel_to_dict(file.read())
        elif filename.endswith(".json"):
            data = json.loads(file.read().decode("utf-8"))
        else:
            return jsonify({"error": "支持 .xlsx 或 .json 格式"}), 400
    except Exception as e:
        return jsonify({"error": f"文件解析失败: {e}"}), 400

    if "lesson" not in data or "sentences" not in data:
        return jsonify({"error": "缺少课程信息或句子数据"}), 400

    if not data["sentences"]:
        return jsonify({"error": "课文句子不能为空"}), 400

    try:
        lesson_id = import_lesson_from_json(data)
    except Exception as e:
        return jsonify({"error": f"导入失败: {e}"}), 500

    # Return preview data
    conn = get_connection()
    full = get_full_lesson_data(conn, lesson_id)
    conn.close()

    settings = data.get("settings", {})

    return jsonify({
        "lesson_id": lesson_id,
        "title_zh": full["lesson"].title_zh,
        "title_en": full["lesson"].title_en,
        "hsk_level": full["lesson"].hsk_level,
        "sentence_count": len(full["sentences"]),
        "vocab_count": len(full["vocabulary"]),
        "grammar_count": len(full["grammar"]),
        "voice": settings.get("voice", "zh-CN-XiaoxiaoNeural"),
        "video_format": settings.get("video_format", "both"),
    })


@app.route("/api/generate", methods=["POST"])
def generate_video():
    """Generate video for a lesson. Returns output paths."""
    import time, traceback
    data = request.get_json()
    lesson_id = data.get("lesson_id")
    voice = data.get("voice", "zh-CN-XiaoxiaoNeural")
    video_format = data.get("video_format", "both")

    if not lesson_id:
        return jsonify({"error": "lesson_id required"}), 400

    output_dir = OUTPUT_DIR / f"lesson_{lesson_id}"
    output_dir.mkdir(parents=True, exist_ok=True)

    t0 = time.time()
    app.logger.info(f"Starting video generation: lesson_id={lesson_id}, voice={voice}, format={video_format}")

    try:
        from scripts.daily_runner import run_pipeline
        success = asyncio.run(
            run_pipeline(lesson_id, output_dir, voice, video_format)
        )
    except Exception as e:
        app.logger.error(f"Generation error: {traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500

    elapsed = round(time.time() - t0, 1)
    app.logger.info(f"Generation finished in {elapsed}s, success={success}")

    if not success:
        return jsonify({"error": "视频生成失败，请查看服务器日志"}), 500

    # Collect output files
    videos = []
    for f in output_dir.glob("final_*.mp4"):
        videos.append({
            "name": f.name,
            "size_mb": round(f.stat().st_size / (1024 * 1024), 1),
            "url": f"/api/download/{lesson_id}/{f.name}",
        })

    return jsonify({"success": True, "lesson_id": lesson_id, "videos": videos})


@app.route("/api/download/<int:lesson_id>/<filename>")
def download_video(lesson_id, filename):
    """Download a generated video file."""
    output_dir = OUTPUT_DIR / f"lesson_{lesson_id}"
    filepath = output_dir / filename
    if not filepath.exists():
        return jsonify({"error": "File not found"}), 404
    return send_file(str(filepath), as_attachment=True, download_name=filename)


@app.route("/api/preview/<int:lesson_id>")
def preview_lesson(lesson_id):
    """Get lesson data for preview."""
    conn = get_connection()
    full = get_full_lesson_data(conn, lesson_id)
    conn.close()
    if not full:
        return jsonify({"error": "Lesson not found"}), 404

    lesson = full["lesson"]
    return jsonify({
        "lesson": {
            "title_zh": lesson.title_zh,
            "title_pinyin": lesson.title_pinyin,
            "title_en": lesson.title_en,
            "hsk_level": lesson.hsk_level,
            "category": lesson.category,
            "business_tip_zh": lesson.business_tip_zh,
            "business_tip_en": lesson.business_tip_en,
        },
        "sentences": [
            {"sort_order": s.sort_order, "text_zh": s.text_zh,
             "text_pinyin": s.text_pinyin, "text_en": s.text_en,
             "highlight_words": s.highlight_words}
            for s in full["sentences"]
        ],
        "vocabulary": [
            {"word_zh": v.word_zh, "word_pinyin": v.word_pinyin,
             "word_en": v.word_en, "word_pos": v.word_pos,
             "hsk_level": v.hsk_level, "icon_emoji": v.icon_emoji,
             "first_appear_sentence": v.first_appear_sentence}
            for v in full["vocabulary"]
        ],
        "grammar": [
            {"pattern_zh": g.pattern_zh, "pattern_pinyin": g.pattern_pinyin,
             "pattern_en": g.pattern_en, "hsk_level": g.hsk_level,
             "first_appear_sentence": g.first_appear_sentence}
            for g in full["grammar"]
        ],
    })


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
