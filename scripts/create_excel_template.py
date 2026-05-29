#!/usr/bin/env python3
"""Create Excel lesson template with sample data."""

import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from scripts.utils.db_helper import TEMPLATES_DIR

BLUE = "2563EB"
DARK = "0F172A"
GOLD = "F59E0B"
WHITE = "FFFFFF"
GRAY = "94A3B8"

header_font = Font(bold=True, color=WHITE, size=12)
header_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
gold_font = Font(bold=True, color=GOLD, size=11)
thin_border = Border(
    left=Side(style="thin", color=GRAY),
    right=Side(style="thin", color=GRAY),
    top=Side(style="thin", color=GRAY),
    bottom=Side(style="thin", color=GRAY),
)


def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border


def style_data(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def create_template(output_path: Path, with_sample: bool = True):
    wb = Workbook()

    # --- Sheet 1: 课程信息 ---
    ws1 = wb.active
    ws1.title = "课程信息"
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 50

    fields = [
        ("字段", "填写内容"),
        ("课程标题（中文）", "个人介绍" if with_sample else ""),
        ("课程标题（拼音）", "gèrén jièshào" if with_sample else ""),
        ("课程标题（英文）", "Self Introduction" if with_sample else ""),
        ("HSK等级", "HSK1-2" if with_sample else "HSK1"),
        ("分类", "商务汉语" if with_sample else "商务汉语"),
        ("商务小贴士（中文）", "在商务场合，自我介绍要简洁、自信、突出核心能力。" if with_sample else ""),
        ("商务小贴士（英文）", "In business situations, self-introduction should be concise, confident and highlight your core strengths." if with_sample else ""),
        ("语音", "zh-CN-XiaoxiaoNeural" if with_sample else "zh-CN-XiaoxiaoNeural"),
        ("视频格式", "both" if with_sample else "both"),
    ]

    for i, (field, value) in enumerate(fields, 1):
        ws1.cell(row=i, column=1, value=field)
        ws1.cell(row=i, column=2, value=value)
        if i == 1:
            style_header(ws1, i, 2)
        else:
            ws1.cell(row=i, column=1).font = Font(bold=True)
            style_data(ws1, i, 2)

    # --- Sheet 2: 课文句子 ---
    ws2 = wb.create_sheet("课文句子")
    headers2 = ["序号", "中文", "拼音", "英文", "高亮词（逗号分隔）"]
    widths2 = [8, 30, 35, 40, 25]
    for i, (h, w) in enumerate(zip(headers2, widths2), 1):
        ws2.cell(row=1, column=i, value=h)
        ws2.column_dimensions[chr(64 + i)].width = w
    style_header(ws2, 1, len(headers2))

    if with_sample:
        sentences = [
            (1, "大家好，我叫张三。", "Dàjiā hǎo, wǒ jiào Zhāng Sān.", "Hello everyone, my name is Zhang San.", "大家,我叫"),
            (2, "我今年二十五岁。", "Wǒ jīnnián èrshíwǔ suì.", "I am twenty-five years old this year.", "今年,岁"),
            (3, "我是中国人。", "Wǒ shì Zhōngguó rén.", "I am Chinese.", "中国"),
            (4, "我在北京工作。", "Wǒ zài Běijīng gōngzuò.", "I work in Beijing.", "北京,工作"),
            (5, "我在一个公司工作。", "Wǒ zài yí gè gōngsī gōngzuò.", "I work at a company.", "公司,工作"),
            (6, "我是销售。", "Wǒ shì xiāoshòu.", "I am in sales.", "销售"),
            (7, "很高兴认识大家。", "Hěn gāoxìng rènshi dàjiā.", "Nice to meet everyone.", "高兴,认识"),
        ]
        for i, row in enumerate(sentences, 2):
            for j, val in enumerate(row, 1):
                ws2.cell(row=i, column=j, value=val)
            style_data(ws2, i, len(headers2))

    # --- Sheet 3: 词汇 ---
    ws3 = wb.create_sheet("词汇")
    headers3 = ["中文", "拼音", "英文", "词性", "HSK等级", "Emoji", "首次出现句号"]
    widths3 = [12, 15, 18, 10, 10, 8, 15]
    for i, (h, w) in enumerate(zip(headers3, widths3), 1):
        ws3.cell(row=1, column=i, value=h)
        ws3.column_dimensions[chr(64 + i)].width = w
    style_header(ws3, 1, len(headers3))

    if with_sample:
        vocab = [
            ("大家", "dàjiā", "everyone", "pron.", "HSK1", "👥", 1),
            ("我叫", "wǒ jiào", "my name is", "v.", "HSK1", "🙋", 1),
            ("中国", "Zhōngguó", "China", "n.", "HSK1", "🇨🇳", 3),
            ("工作", "gōngzuò", "work/job", "n./v.", "HSK1", "💼", 4),
            ("公司", "gōngsī", "company", "n.", "HSK2", "🏢", 5),
            ("销售", "xiāoshòu", "sales", "n./v.", "HSK2", "📊", 6),
            ("高兴", "gāoxìng", "happy/glad", "adj.", "HSK1", "😊", 7),
            ("认识", "rènshi", "to know/meet", "v.", "HSK1", "🤝", 7),
        ]
        for i, row in enumerate(vocab, 2):
            for j, val in enumerate(row, 1):
                ws3.cell(row=i, column=j, value=val)
            style_data(ws3, i, len(headers3))

    # --- Sheet 4: 语法 ---
    ws4 = wb.create_sheet("语法")
    headers4 = ["语法模式", "拼音", "英文", "中文解释", "英文解释", "例句中文", "例句拼音", "例句英文", "HSK等级", "首次出现句号"]
    widths4 = [15, 20, 18, 25, 30, 25, 25, 30, 10, 15]
    for i, (h, w) in enumerate(zip(headers4, widths4), 1):
        ws4.cell(row=1, column=i, value=h)
        col_letter = chr(64 + i) if i <= 26 else chr(64 + (i - 1) // 26) + chr(64 + (i - 1) % 26 + 1)
        ws4.column_dimensions[col_letter].width = w
    style_header(ws4, 1, len(headers4))

    if with_sample:
        grammar = [
            ("我叫……", "wǒ jiào...", "My name is...", "用于自我介绍时说出自己的名字。", "Used to state your name during self-introduction.", "我叫李明。", "Wǒ jiào Lǐ Míng.", "My name is Li Ming.", "HSK1", 1),
            ("我在……工作", "wǒ zài... gōngzuò", "I work at/in...", "表示在某个地点或机构工作。", "Indicates working at a place or organization.", "我在上海一家银行工作。", "Wǒ zài Shànghǎi yì jiā yínháng gōngzuò.", "I work at a bank in Shanghai.", "HSK2", 4),
            ("很高兴认识……", "hěn gāoxìng rènshi...", "Nice to meet...", "初次见面时的礼貌用语。", "A polite expression used when meeting someone for the first time.", "很高兴认识你！", "Hěn gāoxìng rènshi nǐ!", "Nice to meet you!", "HSK1-2", 7),
        ]
        for i, row in enumerate(grammar, 2):
            for j, val in enumerate(row, 1):
                ws4.cell(row=i, column=j, value=val)
            style_data(ws4, i, len(headers4))

    wb.save(output_path)
    print(f"Excel template saved: {output_path}")


if __name__ == "__main__":
    # Create both blank and sample templates
    create_template(TEMPLATES_DIR / "lesson_template.xlsx", with_sample=False)
    create_template(TEMPLATES_DIR / "lesson_sample.xlsx", with_sample=True)
